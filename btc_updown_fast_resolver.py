import argparse
import csv
import json
import re
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None


GAMMA_EVENTS_URL = "https://gamma-api.polymarket.com/events"
CHAINLINK_STREAM_PAGE = "https://data.chain.link/streams/{stream_slug}"
CHAINLINK_LIVE_URL = "https://data.chain.link/api/live-data-engine-stream-data"
CHAINLINK_HISTORICAL_URL = "https://data.chain.link/api/historical-data-engine-stream-data"
SCRIPT_DIR = Path(__file__).resolve().parent
ASSET_5M_CONFIG = {
    "btc": {
        "stream_slug": "btc-usd",
        "market_prefix": "btc-updown-5m-",
        "folder_name": "btc 5min",
        "file_prefix": "btc_updown",
    },
    "eth": {
        "stream_slug": "eth-usd",
        "market_prefix": "eth-updown-5m-",
        "folder_name": "eth 5min",
        "file_prefix": "eth_updown",
    },
    "sol": {
        "stream_slug": "sol-usd",
        "market_prefix": "sol-updown-5m-",
        "folder_name": "sol 5min",
        "file_prefix": "sol_updown",
    },
}


def normalize_asset_name(value: Optional[str]) -> str:
    return str(value or "").strip().lower()


def resolve_asset_config(asset_name: str) -> Dict[str, str]:
    asset = normalize_asset_name(asset_name)
    cfg = ASSET_5M_CONFIG.get(asset)
    if cfg is None:
        supported = ", ".join(sorted(ASSET_5M_CONFIG.keys()))
        raise ValueError(f"Ativo invalido: '{asset_name}'. Use um de: {supported}.")
    return {
        "asset": asset,
        "stream_slug": str(cfg["stream_slug"]),
        "market_prefix": str(cfg["market_prefix"]),
        "folder_name": str(cfg["folder_name"]),
        "file_prefix": str(cfg["file_prefix"]),
    }


def resolve_output_paths(
    asset_cfg: Dict[str, str],
    output_dir: Optional[str],
    output_csv: Optional[str],
    output_xlsx: Optional[str],
    paper_output_csv: Optional[str],
    paper_output_xlsx: Optional[str],
) -> Tuple[str, str, str, str, Path]:
    base_dir = Path(output_dir) if output_dir else (SCRIPT_DIR / asset_cfg["folder_name"])
    file_prefix = asset_cfg["file_prefix"]

    resolved_output_csv = str(Path(output_csv)) if output_csv else str(base_dir / f"{file_prefix}_fast_capture.csv")
    resolved_output_xlsx = str(Path(output_xlsx)) if output_xlsx else str(base_dir / f"{file_prefix}_fast_capture.xlsx")
    resolved_paper_csv = (
        str(Path(paper_output_csv))
        if paper_output_csv
        else str(base_dir / f"{file_prefix}_paper_trade.csv")
    )
    resolved_paper_xlsx = (
        str(Path(paper_output_xlsx))
        if paper_output_xlsx
        else str(base_dir / f"{file_prefix}_paper_trade.xlsx")
    )
    return resolved_output_csv, resolved_output_xlsx, resolved_paper_csv, resolved_paper_xlsx, base_dir


def iso_to_epoch(value: Optional[str]) -> Optional[float]:
    if not value:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    try:
        return datetime.fromisoformat(text).timestamp()
    except Exception:
        return None


def epoch_to_utc_text(ts: Optional[float]) -> str:
    if ts is None:
        return "N/A"
    return datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S.%f")[:-3] + "Z"


def parse_next_data_payload(html: str) -> dict:
    match = re.search(
        r'<script id="__NEXT_DATA__" type="application/json">(.*?)</script>',
        html,
        flags=re.S,
    )
    if not match:
        raise RuntimeError("Nao foi possivel localizar __NEXT_DATA__ na pagina da Chainlink.")
    return json.loads(match.group(1))


def get_primary_abi_index(schema_name: str) -> int:
    schema = str(schema_name or "").strip()
    mapping = {
        "Premium": 0,
        "v7": 0,
        "v8": 1,
        "v9": 0,
        "v10": 1,
        "v11": 0,
    }
    return mapping.get(schema, 0)


def resolve_chainlink_stream(session: requests.Session, stream_slug: str, timeout: float) -> Tuple[str, int, str]:
    url = CHAINLINK_STREAM_PAGE.format(stream_slug=stream_slug)
    resp = session.get(url, timeout=timeout)
    resp.raise_for_status()
    payload = parse_next_data_payload(resp.text)

    stream_data = (((payload.get("props") or {}).get("pageProps") or {}).get("streamData") or {})
    meta = stream_data.get("streamMetadata") or {}
    docs = meta.get("docs") or {}
    schema = docs.get("schema") or "Premium"
    abi_index = get_primary_abi_index(schema)

    extra_cfg = stream_data.get("extraConfig") or {}
    feed_id = extra_cfg.get("feedId") or meta.get("feedId")
    if not feed_id:
        raise RuntimeError("Nao foi possivel obter feedId da Chainlink.")

    return str(feed_id), int(abi_index), str(schema)


def fetch_market_by_slug(
    session: requests.Session,
    slug: str,
    timeout: float,
) -> Optional[dict]:
    resp = session.get(GAMMA_EVENTS_URL, params={"slug": slug}, timeout=timeout)
    resp.raise_for_status()
    data = resp.json()
    if not isinstance(data, list) or not data:
        return None

    event = data[0]
    markets = event.get("markets", [])
    if not markets:
        return None
    return markets[0]


def fetch_chainlink_nodes(
    session: requests.Session,
    feed_id: str,
    abi_index: int,
    timeout: float,
) -> List[dict]:
    params = {
        "feedId": feed_id,
        "abiIndex": str(abi_index),
        "queryWindow": "1m",
    }
    resp = session.get(CHAINLINK_LIVE_URL, params=params, timeout=timeout)
    resp.raise_for_status()
    data = resp.json()
    return (((data.get("data") or {}).get("allStreamValuesGenerics") or {}).get("nodes") or [])


def fetch_chainlink_historical_nodes(
    session: requests.Session,
    feed_id: str,
    abi_index: int,
    timeout: float,
) -> List[dict]:
    params = {
        "feedId": feed_id,
        "abiIndex": str(abi_index),
        "timeRange": "1D",
    }
    resp = session.get(CHAINLINK_HISTORICAL_URL, params=params, timeout=timeout)
    resp.raise_for_status()
    data = resp.json()
    return (((data.get("data") or {}).get("allStreamValuesGeneric1Minutes") or {}).get("nodes") or [])


def parse_candlestick_price(candlestick: str, field_name: str) -> Tuple[Optional[float], Optional[float]]:
    """
    Extrai timestamp e valor de um campo do candlestick da Chainlink.
    Exemplo de campo:
    open:(ts:"2026-04-13 12:08:00.339431+00",val:70948.5761)
    """
    text = str(candlestick or "")
    pattern = rf'{re.escape(field_name)}:\(ts:"([^"]+)",val:([-\d.]+)\)'
    match = re.search(pattern, text)
    if not match:
        return None, None

    ts_raw = match.group(1).strip()
    # "2026-04-13 12:08:00.339431+00" -> "2026-04-13T12:08:00.339431+00:00"
    ts_norm = ts_raw.replace(" ", "T", 1)
    if re.search(r"[+-]\d{2}$", ts_norm):
        ts_norm = ts_norm + ":00"

    ts_epoch = iso_to_epoch(ts_norm)
    try:
        value = float(match.group(2))
    except Exception:
        value = None

    return ts_epoch, value


def parse_list_field(raw_value) -> List:
    if isinstance(raw_value, list):
        return raw_value
    if isinstance(raw_value, str):
        text = raw_value.strip()
        if not text:
            return []
        try:
            parsed = json.loads(text)
            if isinstance(parsed, list):
                return parsed
        except json.JSONDecodeError:
            return []
    return []


def normalize_outcome_name(name: Optional[str]) -> Optional[str]:
    if name is None:
        return None
    text = str(name).strip().upper()
    if text == "UP":
        return "UP"
    if text == "DOWN":
        return "DOWN"
    return str(name).strip()


def infer_official_outcome(outcomes_raw, outcome_prices_raw) -> Optional[str]:
    outcomes = [str(item) for item in parse_list_field(outcomes_raw)]
    prices_raw = parse_list_field(outcome_prices_raw)
    prices: List[float] = []
    for item in prices_raw:
        try:
            prices.append(float(item))
        except Exception:
            prices.append(0.0)

    if not outcomes or not prices:
        return None

    winner_idx = None
    for idx, price in enumerate(prices):
        if price >= 0.999999:
            winner_idx = idx
            break

    if winner_idx is None:
        return None
    if winner_idx >= len(outcomes):
        return None

    return normalize_outcome_name(outcomes[winner_idx])


def extract_up_down_prices(market: dict) -> Tuple[Optional[float], Optional[float]]:
    outcomes = [str(x).strip().upper() for x in parse_list_field(market.get("outcomes"))]
    prices_raw = parse_list_field(market.get("outcomePrices"))
    prices: List[Optional[float]] = []
    for item in prices_raw:
        try:
            prices.append(float(item))
        except Exception:
            prices.append(None)

    side_to_price: Dict[str, float] = {}
    for idx, side in enumerate(outcomes):
        if idx >= len(prices):
            continue
        price = prices[idx]
        if price is None:
            continue
        if side in ("UP", "DOWN"):
            side_to_price[side] = price

    # Fallback seguro para mercados Up/Down: ordem costuma ser [Up, Down].
    if "UP" not in side_to_price and len(prices) >= 1 and prices[0] is not None:
        side_to_price["UP"] = float(prices[0])
    if "DOWN" not in side_to_price and len(prices) >= 2 and prices[1] is not None:
        side_to_price["DOWN"] = float(prices[1])

    return side_to_price.get("UP"), side_to_price.get("DOWN")


def get_price_from_attribute_map(attr_map: Dict[str, float]) -> Tuple[Optional[float], str]:
    # Prioridade para o benchmark usado no stream de referência.
    for key in ("benchmark", "price", "mid"):
        if key in attr_map:
            return attr_map[key], key

    if "bid" in attr_map and "ask" in attr_map:
        return (attr_map["bid"] + attr_map["ask"]) / 2.0, "mid(bid+ask)"

    if attr_map:
        k = next(iter(attr_map.keys()))
        return attr_map[k], k

    return None, "none"


@dataclass
class CycleState:
    slug: str
    question: str
    start_ts: float
    end_ts: float
    up_price: Optional[float] = None
    down_price: Optional[float] = None
    target_price: Optional[float] = None
    target_ts: Optional[float] = None
    target_src: Optional[str] = None
    close_price: Optional[float] = None
    close_ts: Optional[float] = None
    close_src: Optional[str] = None
    chainlink_result: Optional[str] = None
    official_result: Optional[str] = None
    official_candidate_result: Optional[str] = None
    official_candidate_hits: int = 0
    resolution_checked_at: Optional[float] = None
    inferred_result: Optional[str] = None
    warned_target_delay: bool = False
    warned_close_wait: bool = False
    warned_resolution_wait: bool = False
    warned_resolution_fallback: bool = False


@dataclass
class PaperPosition:
    slug: str
    question: str
    start_ts: float
    end_ts: float
    side: str
    shares: float
    entry_price: float
    entry_ts: float
    entry_cost: float
    sequence_index_before: int
    gale_factor: float


class ResultStore:
    HEADERS = [
        "captured_at_utc",
        "slug",
        "question",
        "start_time_utc",
        "end_time_utc",
        "target_price",
        "target_ts_utc",
        "target_src",
        "close_price",
        "close_ts_utc",
        "close_src",
        "delta",
        "inferred_result",
        "decision_delay_seconds",
    ]

    def __init__(
        self,
        csv_path: str,
        xlsx_path: Optional[str],
        enable_xlsx: bool,
    ) -> None:
        self.csv_path = Path(csv_path)
        self.csv_path.parent.mkdir(parents=True, exist_ok=True)

        self.enable_xlsx = bool(enable_xlsx and Workbook is not None and load_workbook is not None)
        self.xlsx_path = Path(xlsx_path) if xlsx_path else None
        self._xlsx_disabled_reason: Optional[str] = None
        if self.enable_xlsx and self.xlsx_path is not None:
            self.xlsx_path.parent.mkdir(parents=True, exist_ok=True)

        self._seen_slugs = self._load_existing_slugs_from_csv()
        self._ensure_csv_header()
        self._ensure_xlsx_header_if_needed()

        if enable_xlsx and not self.enable_xlsx:
            print("[AVISO] openpyxl nao disponivel; exportando apenas CSV.")

    def _load_existing_slugs_from_csv(self) -> set:
        slugs = set()
        if not self.csv_path.exists():
            return slugs

        try:
            with self.csv_path.open("r", newline="", encoding="utf-8") as fh:
                reader = csv.DictReader(fh)
                for row in reader:
                    slug = str(row.get("slug") or "").strip()
                    if slug:
                        slugs.add(slug)
        except Exception:
            # Se arquivo estiver corrompido/parcial, seguimos sem bloqueio.
            pass
        return slugs

    def _ensure_csv_header(self) -> None:
        if self.csv_path.exists() and self.csv_path.stat().st_size > 0:
            return
        with self.csv_path.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=self.HEADERS)
            writer.writeheader()

    def _ensure_xlsx_header_if_needed(self) -> None:
        if not self.enable_xlsx or self.xlsx_path is None:
            return

        if self.xlsx_path.exists():
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "captura"
        ws.append(self.HEADERS)
        wb.save(self.xlsx_path)

    def _append_csv(self, row: Dict[str, str]) -> None:
        with self.csv_path.open("a", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=self.HEADERS)
            writer.writerow(row)

    def _append_xlsx(self, row: Dict[str, str]) -> None:
        if not self.enable_xlsx or self.xlsx_path is None:
            return
        wb = load_workbook(self.xlsx_path)
        ws = wb["captura"] if "captura" in wb.sheetnames else wb.active
        ws.append([row.get(col, "") for col in self.HEADERS])
        wb.save(self.xlsx_path)

    def save_inference(self, cycle: CycleState, delta: Optional[float], decision_delay_seconds: float) -> None:
        if cycle.slug in self._seen_slugs:
            return

        captured_at = epoch_to_utc_text(time.time())
        row = {
            "captured_at_utc": captured_at,
            "slug": cycle.slug,
            "question": cycle.question,
            "start_time_utc": epoch_to_utc_text(cycle.start_ts),
            "end_time_utc": epoch_to_utc_text(cycle.end_ts),
            "target_price": f"{cycle.target_price:.10f}" if cycle.target_price is not None else "",
            "target_ts_utc": epoch_to_utc_text(cycle.target_ts),
            "target_src": str(cycle.target_src or ""),
            "close_price": f"{cycle.close_price:.10f}" if cycle.close_price is not None else "",
            "close_ts_utc": epoch_to_utc_text(cycle.close_ts),
            "close_src": str(cycle.close_src or ""),
            "delta": f"{delta:+.10f}" if delta is not None else "",
            "inferred_result": str(cycle.inferred_result or ""),
            "decision_delay_seconds": f"{decision_delay_seconds:.2f}",
        }

        wrote_any = False
        try:
            self._append_csv(row)
            wrote_any = True
        except Exception as e:
            print(f"[ERRO] Falha ao salvar inferencia no CSV ({self.csv_path}): {e}")

        if self.enable_xlsx:
            try:
                self._append_xlsx(row)
                wrote_any = True
            except Exception as e:
                self.enable_xlsx = False
                self._xlsx_disabled_reason = str(e)
                print(
                    f"[AVISO] XLSX de inferencias indisponivel ({self.xlsx_path}): {e}. "
                    "Provavel arquivo aberto/bloqueado. Seguindo com CSV."
                )

        if wrote_any:
            self._seen_slugs.add(cycle.slug)


class PaperTradeStore:
    HEADERS = [
        "captured_at_utc",
        "slug",
        "question",
        "start_time_utc",
        "end_time_utc",
        "side",
        "shares",
        "entry_price",
        "entry_time_utc",
        "entry_cost",
        "resolved_result",
        "won",
        "payout",
        "trade_pnl",
        "cash_after",
        "pnl_after",
        "sequence_index_before",
        "sequence_index_after",
        "sequence_pattern",
    ]

    def __init__(self, csv_path: str, xlsx_path: Optional[str], enable_xlsx: bool) -> None:
        self.csv_path = Path(csv_path)
        self.csv_path.parent.mkdir(parents=True, exist_ok=True)

        self.enable_xlsx = bool(enable_xlsx and Workbook is not None and load_workbook is not None)
        self.xlsx_path = Path(xlsx_path) if xlsx_path else None
        self._xlsx_disabled_reason: Optional[str] = None
        if self.enable_xlsx and self.xlsx_path is not None:
            self.xlsx_path.parent.mkdir(parents=True, exist_ok=True)

        self._seen_slugs = self._load_existing_slugs_from_csv()
        self._ensure_csv_header()
        self._ensure_xlsx_header_if_needed()

    def _load_existing_slugs_from_csv(self) -> set:
        slugs = set()
        if not self.csv_path.exists():
            return slugs

        try:
            with self.csv_path.open("r", newline="", encoding="utf-8") as fh:
                reader = csv.DictReader(fh)
                for row in reader:
                    slug = str(row.get("slug") or "").strip()
                    if slug:
                        slugs.add(slug)
        except Exception:
            pass
        return slugs

    def _ensure_csv_header(self) -> None:
        if self.csv_path.exists() and self.csv_path.stat().st_size > 0:
            return
        with self.csv_path.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=self.HEADERS)
            writer.writeheader()

    def _ensure_xlsx_header_if_needed(self) -> None:
        if not self.enable_xlsx or self.xlsx_path is None:
            return
        if self.xlsx_path.exists():
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "paper_trade"
        ws.append(self.HEADERS)
        wb.save(self.xlsx_path)

    def _append_csv(self, row: Dict[str, str]) -> None:
        with self.csv_path.open("a", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=self.HEADERS)
            writer.writerow(row)

    def _append_xlsx(self, row: Dict[str, str]) -> None:
        if not self.enable_xlsx or self.xlsx_path is None:
            return
        wb = load_workbook(self.xlsx_path)
        ws = wb["paper_trade"] if "paper_trade" in wb.sheetnames else wb.active
        ws.append([row.get(col, "") for col in self.HEADERS])
        wb.save(self.xlsx_path)

    def save_settlement(
        self,
        position: PaperPosition,
        resolved_result: str,
        won: bool,
        payout: float,
        trade_pnl: float,
        cash_after: float,
        pnl_after: float,
        sequence_index_after: int,
        sequence_pattern: List[str],
    ) -> None:
        if position.slug in self._seen_slugs:
            return

        row = {
            "captured_at_utc": epoch_to_utc_text(time.time()),
            "slug": position.slug,
            "question": position.question,
            "start_time_utc": epoch_to_utc_text(position.start_ts),
            "end_time_utc": epoch_to_utc_text(position.end_ts),
            "side": position.side,
            "shares": f"{position.shares:.4f}",
            "entry_price": f"{position.entry_price:.10f}",
            "entry_time_utc": epoch_to_utc_text(position.entry_ts),
            "entry_cost": f"{position.entry_cost:.10f}",
            "resolved_result": resolved_result,
            "won": "1" if won else "0",
            "payout": f"{payout:.10f}",
            "trade_pnl": f"{trade_pnl:+.10f}",
            "cash_after": f"{cash_after:.10f}",
            "pnl_after": f"{pnl_after:+.10f}",
            "sequence_index_before": str(position.sequence_index_before),
            "sequence_index_after": str(sequence_index_after),
            "sequence_pattern": ",".join(sequence_pattern),
        }

        wrote_any = False
        try:
            self._append_csv(row)
            wrote_any = True
        except Exception as e:
            print(f"[ERRO] Falha ao salvar paper trade no CSV ({self.csv_path}): {e}")

        if self.enable_xlsx:
            try:
                self._append_xlsx(row)
                wrote_any = True
            except Exception as e:
                self.enable_xlsx = False
                self._xlsx_disabled_reason = str(e)
                print(
                    f"[AVISO] XLSX de paper trade indisponivel ({self.xlsx_path}): {e}. "
                    "Provavel arquivo aberto/bloqueado. Seguindo com CSV."
                )

        if wrote_any:
            self._seen_slugs.add(position.slug)


class FastResolver:
    @staticmethod
    def _parse_sequence(sequence_text: str) -> List[str]:
        raw_items = [item.strip().upper() for item in str(sequence_text or "").split(",")]
        items = [item for item in raw_items if item]
        if not items:
            raise ValueError("Sequencia vazia. Exemplo valido: UP,UP,DOWN,DOWN")
        for item in items:
            if item not in ("UP", "DOWN"):
                raise ValueError(f"Lado invalido na sequencia: {item}. Use apenas UP ou DOWN.")
        return items

    def __init__(
        self,
        asset_name: str,
        stream_slug: str,
        market_prefix: str,
        interval_seconds: float,
        decision_delay_seconds: float,
        timeout_seconds: float,
        require_official_resolution: bool,
        allow_chainlink_fallback: bool,
        use_official_ws_resolution: bool,
        official_resolution_wait_seconds: float,
        official_resolution_poll_seconds: float,
        official_confirmation_polls: int,
        output_csv: str,
        output_xlsx: Optional[str],
        enable_xlsx: bool,
        paper_starting_cash: float,
        paper_shares: float,
        paper_stake_usd: float,
        paper_sequence: str,
        paper_entry_cutoff_seconds: float,
        paper_entry_min_price: float,
        paper_entry_max_price: float,
        paper_gale_multiplier: float,
        paper_gale_max_steps: int,
        paper_output_csv: str,
        paper_output_xlsx: Optional[str],
        enable_paper_xlsx: bool,
    ) -> None:
        self.asset_name = normalize_asset_name(asset_name)
        self.stream_slug = stream_slug
        self.market_prefix = market_prefix
        self.interval_seconds = interval_seconds
        self.decision_delay_seconds = decision_delay_seconds
        self.timeout_seconds = timeout_seconds
        self.require_official_resolution = bool(require_official_resolution)
        self.allow_chainlink_fallback = bool(allow_chainlink_fallback)
        self.use_official_ws_resolution = bool(use_official_ws_resolution)
        self.official_resolution_wait_seconds = max(0.0, float(official_resolution_wait_seconds))
        self.official_resolution_poll_seconds = max(0.2, float(official_resolution_poll_seconds))
        self.official_confirmation_polls = max(1, int(official_confirmation_polls))
        self.result_store = ResultStore(
            csv_path=output_csv,
            xlsx_path=output_xlsx,
            enable_xlsx=enable_xlsx,
        )
        self.paper_store = PaperTradeStore(
            csv_path=paper_output_csv,
            xlsx_path=paper_output_xlsx,
            enable_xlsx=enable_paper_xlsx,
        )
        self.paper_sequence = self._parse_sequence(paper_sequence)
        self.paper_sequence_idx = 0
        self.paper_boot_ts = time.time()
        self.paper_starting_cash = max(0.0, float(paper_starting_cash))
        self.paper_cash = self.paper_starting_cash
        self.paper_shares = max(0.01, float(paper_shares))
        self.paper_stake_usd = max(0.0, float(paper_stake_usd))
        self.paper_sizing_mode = "USD" if self.paper_stake_usd > 0 else "SHARES"
        self.paper_entry_cutoff_seconds = max(0.5, float(paper_entry_cutoff_seconds))
        min_price = max(0.0, float(paper_entry_min_price))
        max_price = max(0.0, float(paper_entry_max_price))
        if max_price < min_price:
            min_price, max_price = max_price, min_price
        self.paper_entry_min_price = min_price
        self.paper_entry_max_price = max_price
        self.paper_gale_multiplier = max(1.0, float(paper_gale_multiplier))
        self.paper_gale_max_steps = max(0, int(paper_gale_max_steps))
        self.paper_loss_streak = 0
        self.paper_open_position: Optional[PaperPosition] = None
        self.paper_entered_slugs = set()
        self.paper_skipped_slugs = set()
        self.paper_ignored_slugs = set()
        self.paper_range_warned_keys = set()
        self.paper_trades = 0
        self.paper_wins = 0
        self.paper_losses = 0

        self.session = requests.Session()
        self.feed_id, self.abi_index, self.schema = resolve_chainlink_stream(
            self.session, self.stream_slug, self.timeout_seconds
        )

        # timestamp -> price (benchmark/mid)
        self.price_by_ts: Dict[float, float] = {}
        self.price_src_by_ts: Dict[float, str] = {}

        # bucket_minute_ts -> (open_ts, open_price, close_ts, close_price)
        self.historical_by_bucket: Dict[int, Tuple[Optional[float], Optional[float], Optional[float], Optional[float]]] = {}
        self.historical_loaded_at: float = 0.0

        # start_ts -> cycle data
        self.cycles: Dict[int, CycleState] = {}

        self._ws_resolver = None
        self._ws_enabled = False
        self._init_official_ws_resolver()

    def build_slug(self, start_ts: int) -> str:
        return f"{self.market_prefix}{start_ts}"

    def _init_official_ws_resolver(self) -> None:
        if not self.use_official_ws_resolution:
            return
        try:
            import polymarket_resolver as ws_resolver  # type: ignore
        except Exception as e:
            print(f"[AVISO] WebSocket resolver oficial indisponivel (import): {e}")
            self._ws_resolver = None
            self._ws_enabled = False
            return

        self._ws_resolver = ws_resolver
        try:
            started = bool(ws_resolver.start_resolver())
        except Exception as e:
            print(f"[AVISO] Falha ao iniciar WebSocket resolver oficial: {e}")
            started = False

        self._ws_enabled = bool(started and ws_resolver.is_resolver_running())
        if self._ws_enabled:
            print("[INIT] official_ws=ON")
        else:
            print("[AVISO] official_ws=OFF (seguindo com polling oficial via Gamma)")

    def _prime_official_ws_slug(self, slug: str) -> None:
        if not self._ws_enabled or self._ws_resolver is None:
            return
        try:
            self._ws_resolver.ensure_slug_registered(slug)
        except Exception:
            # Nao bloqueia o fluxo principal.
            return

    def _fetch_official_result_from_ws(self, slug: str) -> Optional[str]:
        if not self._ws_enabled or self._ws_resolver is None:
            return None
        try:
            result = self._ws_resolver.peek_result_for_slug(slug)
        except Exception:
            return None
        text = str(result or "").strip().upper()
        if text in ("UP", "DOWN"):
            return text
        return None

    def update_live_price_cache(self) -> None:
        nodes = fetch_chainlink_nodes(
            self.session, self.feed_id, self.abi_index, timeout=self.timeout_seconds
        )

        # Agrupa atributos por timestamp
        grouped: Dict[float, Dict[str, float]] = {}
        for item in nodes:
            try:
                ts = iso_to_epoch(item.get("validAfterTs"))
                if ts is None:
                    continue
                attr = str(item.get("attributeName") or "").strip().lower()
                if not attr:
                    continue
                value = float(item.get("valueNumeric"))
            except Exception:
                continue
            grouped.setdefault(ts, {})[attr] = value

        for ts, attr_map in grouped.items():
            price, src = get_price_from_attribute_map(attr_map)
            if price is None:
                continue
            self.price_by_ts[ts] = price
            self.price_src_by_ts[ts] = src

        # Mantem apenas os ultimos 20 minutos no cache local.
        cutoff = time.time() - 1200.0
        old_keys = [ts for ts in self.price_by_ts if ts < cutoff]
        for ts in old_keys:
            self.price_by_ts.pop(ts, None)
            self.price_src_by_ts.pop(ts, None)

    def refresh_historical_cache_if_needed(self, force: bool = False) -> None:
        now = time.time()
        if not force and (now - self.historical_loaded_at) < 25.0:
            return

        nodes = fetch_chainlink_historical_nodes(
            self.session, self.feed_id, self.abi_index, timeout=self.timeout_seconds
        )
        parsed: Dict[int, Tuple[Optional[float], Optional[float], Optional[float], Optional[float]]] = {}

        for node in nodes:
            if str(node.get("attributeName") or "").strip().lower() != "benchmark":
                continue
            bucket_ts = iso_to_epoch(node.get("bucket"))
            if bucket_ts is None:
                continue
            open_ts, open_price = parse_candlestick_price(node.get("candlestick"), "open")
            close_ts, close_price = parse_candlestick_price(node.get("candlestick"), "close")
            parsed[int(bucket_ts)] = (open_ts, open_price, close_ts, close_price)

        if parsed:
            self.historical_by_bucket = parsed
            self.historical_loaded_at = now

    def get_historical_open_at_minute(self, minute_ts: float) -> Tuple[Optional[float], Optional[float]]:
        bucket = int(minute_ts) - (int(minute_ts) % 60)
        entry = self.historical_by_bucket.get(bucket)
        if not entry:
            return None, None
        open_ts, open_price, _, _ = entry
        return open_ts, open_price

    def _refresh_cycle_side_prices(self, cycle: CycleState) -> None:
        market = fetch_market_by_slug(self.session, cycle.slug, timeout=self.timeout_seconds)
        if not market:
            return
        up_price, down_price = extract_up_down_prices(market)
        if up_price is not None:
            cycle.up_price = up_price
        if down_price is not None:
            cycle.down_price = down_price

    def _maybe_refresh_official_result(self, cycle: CycleState, now_ts: float) -> Optional[str]:
        if cycle.official_result is not None:
            return cycle.official_result

        # Tenta resultado oficial via WebSocket (nao bloqueante).
        ws_result = self._fetch_official_result_from_ws(cycle.slug)
        if ws_result in ("UP", "DOWN"):
            cycle.official_result = ws_result
            return cycle.official_result

        if cycle.resolution_checked_at is not None:
            elapsed = now_ts - cycle.resolution_checked_at
            if elapsed < self.official_resolution_poll_seconds:
                return None

        cycle.resolution_checked_at = now_ts
        # Mantem o slug "primed" no WS para acelerar deteccao quando habilitado.
        self._prime_official_ws_slug(cycle.slug)
        market = fetch_market_by_slug(self.session, cycle.slug, timeout=self.timeout_seconds)
        if not market:
            return None

        official = infer_official_outcome(market.get("outcomes"), market.get("outcomePrices"))
        if official not in ("UP", "DOWN"):
            cycle.official_candidate_result = None
            cycle.official_candidate_hits = 0
            return None

        if cycle.official_candidate_result == official:
            cycle.official_candidate_hits += 1
        else:
            cycle.official_candidate_result = official
            cycle.official_candidate_hits = 1

        if cycle.official_candidate_hits >= self.official_confirmation_polls:
            cycle.official_result = official
        return cycle.official_result

    def _current_gale_steps(self) -> int:
        if self.paper_gale_max_steps > 0:
            return min(self.paper_loss_streak, self.paper_gale_max_steps)
        return max(0, self.paper_loss_streak)

    def _current_gale_factor(self) -> float:
        try:
            return float(self.paper_gale_multiplier ** self._current_gale_steps())
        except OverflowError:
            return float("inf")

    def _current_target_stake_usd(self) -> Optional[float]:
        if self.paper_sizing_mode != "USD":
            return None
        return float(self.paper_stake_usd * self._current_gale_factor())

    def _current_shares_for_price(self, entry_price: float) -> float:
        if entry_price <= 0:
            return 0.0
        gale = self._current_gale_factor()
        if self.paper_sizing_mode == "USD":
            target_cost = self.paper_stake_usd * gale
            return float(target_cost / entry_price) if target_cost > 0 else 0.0
        return float(self.paper_shares * gale)

    def _current_cost_for_price(self, entry_price: float) -> float:
        if entry_price <= 0:
            return 0.0
        gale = self._current_gale_factor()
        if self.paper_sizing_mode == "USD":
            return float(self.paper_stake_usd * gale)
        return float(self.paper_shares * gale * entry_price)

    def _maybe_open_paper_trade(self, now_ts: float) -> None:
        if self.paper_open_position is not None:
            return

        eligible: List[CycleState] = []
        for cycle in self.cycles.values():
            if cycle.slug in self.paper_entered_slugs:
                continue
            if cycle.slug in self.paper_skipped_slugs:
                continue
            if cycle.slug in self.paper_ignored_slugs:
                continue
            if cycle.inferred_result is not None:
                continue
            age = now_ts - cycle.start_ts
            if age < 0:
                continue
            if age > self.paper_entry_cutoff_seconds:
                continue
            eligible.append(cycle)

        if not eligible:
            return

        side = self.paper_sequence[self.paper_sequence_idx]

        # Tenta do ciclo mais recente para o mais antigo ainda elegivel.
        # Assim continuamos tentando ate o fim do cutoff.
        for cycle in sorted(eligible, key=lambda c: c.start_ts, reverse=True):
            # Atualiza preco sempre para tentar com dado mais recente.
            self._refresh_cycle_side_prices(cycle)
            entry_price = cycle.up_price if side == "UP" else cycle.down_price

            if entry_price is None or entry_price <= 0:
                print(f"[PAPER] {cycle.slug} sem preco de entrada para lado {side}; tentativa adiada.")
                continue

            if not (self.paper_entry_min_price <= entry_price <= self.paper_entry_max_price):
                warn_key = f"{cycle.slug}:{side}"
                if warn_key not in self.paper_range_warned_keys:
                    print(
                        f"[PAPER] {cycle.slug} lado {side} fora da faixa permitida "
                        f"({entry_price:.4f} fora de {self.paper_entry_min_price:.2f}-{self.paper_entry_max_price:.2f})."
                    )
                    self.paper_range_warned_keys.add(warn_key)
                continue

            current_shares = self._current_shares_for_price(entry_price)
            current_gale_factor = self._current_gale_factor()
            entry_cost = self._current_cost_for_price(entry_price)
            if entry_cost > (self.paper_cash + 1e-9):
                print(
                    f"[PAPER] Saldo insuficiente para entrada em {cycle.slug}. "
                    f"cash={self.paper_cash:.4f} | cost={entry_cost:.4f}"
                )
                continue

            position = PaperPosition(
                slug=cycle.slug,
                question=cycle.question,
                start_ts=cycle.start_ts,
                end_ts=cycle.end_ts,
                side=side,
                shares=current_shares,
                entry_price=entry_price,
                entry_ts=now_ts,
                entry_cost=entry_cost,
                sequence_index_before=self.paper_sequence_idx,
                gale_factor=current_gale_factor,
            )
            self.paper_open_position = position
            self.paper_cash -= entry_cost
            self.paper_entered_slugs.add(cycle.slug)
            self.paper_range_warned_keys.discard(f"{cycle.slug}:{side}")

            print(
                f"[PAPER ENTRY] {cycle.slug} | side={side} | shares={current_shares:.2f} | "
                f"gale=x{current_gale_factor:.2f} | "
                f"entry={entry_price:.4f} | cost={entry_cost:.4f} | "
                f"mode={self.paper_sizing_mode} | cash_after_entry={self.paper_cash:.4f} | "
                f"seq_step={self.paper_sequence_idx + 1}/{len(self.paper_sequence)}"
            )
            return

    def _advance_sequence_for_missed_entries(self, now_ts: float) -> None:
        """
        Se passou da janela de entrada e não abrimos posição no ciclo, pulamos 1 passo da sequência.
        """
        for cycle in sorted(self.cycles.values(), key=lambda c: c.start_ts):
            if cycle.slug in self.paper_entered_slugs:
                continue
            if cycle.slug in self.paper_skipped_slugs:
                continue
            if cycle.slug in self.paper_ignored_slugs:
                continue

            if self.paper_open_position is not None and self.paper_open_position.slug == cycle.slug:
                continue

            age = now_ts - cycle.start_ts
            if age <= self.paper_entry_cutoff_seconds:
                continue

            # Se o ciclo já estava fora da janela quando o bot iniciou,
            # ignoramos sem avançar sequência (evita "pular duas vezes" no bootstrap).
            if (cycle.start_ts + self.paper_entry_cutoff_seconds) <= self.paper_boot_ts:
                self.paper_ignored_slugs.add(cycle.slug)
                print(
                    f"[PAPER IGNORE] {cycle.slug} | ciclo expirado antes do start do bot; "
                    "sequencia mantida."
                )
                continue

            side = self.paper_sequence[self.paper_sequence_idx]
            self._refresh_cycle_side_prices(cycle)
            side_price = cycle.up_price if side == "UP" else cycle.down_price

            if side_price is None:
                reason = "sem preco de entrada"
            elif side_price < self.paper_entry_min_price or side_price > self.paper_entry_max_price:
                reason = (
                    f"preco fora da faixa {self.paper_entry_min_price:.2f}-{self.paper_entry_max_price:.2f} "
                    f"(preco={side_price:.4f})"
                )
            else:
                needed_shares = self._current_shares_for_price(side_price)
                needed_cost = self._current_cost_for_price(side_price)

            if side_price is not None and side_price >= self.paper_entry_min_price and side_price <= self.paper_entry_max_price and needed_cost > (self.paper_cash + 1e-9):
                reason = (
                    f"saldo insuficiente (cash={self.paper_cash:.4f} | cost={needed_cost:.4f} | shares={needed_shares:.2f} | mode={self.paper_sizing_mode})"
                )
            elif side_price is not None and self.paper_entry_min_price <= side_price <= self.paper_entry_max_price:
                reason = "janela de entrada expirou"

            seq_before = self.paper_sequence_idx
            self.paper_sequence_idx = (self.paper_sequence_idx + 1) % len(self.paper_sequence)
            self.paper_skipped_slugs.add(cycle.slug)

            print(
                f"[PAPER SKIP] {cycle.slug} | side={side} | motivo={reason} | "
                f"seq {seq_before + 1}/{len(self.paper_sequence)} -> "
                f"{self.paper_sequence_idx + 1}/{len(self.paper_sequence)} | "
                f"next_side={self.paper_sequence[self.paper_sequence_idx]}"
            )

    def _settle_paper_trade_if_needed(self, cycle: CycleState) -> None:
        if self.paper_open_position is None:
            return
        position = self.paper_open_position
        if position.slug != cycle.slug:
            return
        if cycle.inferred_result is None:
            return

        resolved_result = str(cycle.inferred_result).upper()
        won = position.side == resolved_result
        payout = position.shares if won else 0.0
        trade_pnl = payout - position.entry_cost
        self.paper_cash += payout
        pnl_after = self.paper_cash - self.paper_starting_cash

        self.paper_trades += 1
        if won:
            self.paper_wins += 1
            self.paper_loss_streak = 0
        else:
            self.paper_losses += 1
            self.paper_loss_streak += 1
        sequence_index_after = (position.sequence_index_before + 1) % len(self.paper_sequence)
        self.paper_sequence_idx = sequence_index_after
        next_gale_factor = self._current_gale_factor()
        if self.paper_sizing_mode == "USD":
            next_size_text = f"next_stake_usd={self._current_target_stake_usd():.4f}"
        else:
            next_shares = self._current_shares_for_price(max(position.entry_price, 1e-9))
            next_size_text = f"next_shares={next_shares:.2f}"

        print(
            f"[PAPER CLOSE] {position.slug} | expected={position.side} | resolved={resolved_result} | "
            f"won={'YES' if won else 'NO'} | trade_pnl={trade_pnl:+.4f} | "
            f"cash={self.paper_cash:.4f} | pnl={pnl_after:+.4f} | "
            f"record={self.paper_wins}W-{self.paper_losses}L | next_side={self.paper_sequence[self.paper_sequence_idx]} | "
            f"next_gale=x{next_gale_factor:.2f} | {next_size_text}"
        )

        # Fecha posicao em memoria antes de persistir para evitar travar sequencia
        # caso a escrita em arquivo falhe temporariamente.
        self.paper_open_position = None
        try:
            self.paper_store.save_settlement(
                position=position,
                resolved_result=resolved_result,
                won=won,
                payout=payout,
                trade_pnl=trade_pnl,
                cash_after=self.paper_cash,
                pnl_after=pnl_after,
                sequence_index_after=sequence_index_after,
                sequence_pattern=self.paper_sequence,
            )
        except Exception as e:
            print(f"[ERRO] Falha ao persistir fechamento do paper trade: {e}")

    def ensure_cycle(self, start_ts: int) -> None:
        if start_ts in self.cycles:
            return

        slug = self.build_slug(start_ts)
        market = fetch_market_by_slug(self.session, slug, timeout=self.timeout_seconds)
        if not market:
            return

        market_start = iso_to_epoch(market.get("eventStartTime"))
        market_end = iso_to_epoch(market.get("endDate"))
        question = str(market.get("question") or slug)

        if market_start is None:
            market_start = float(start_ts)
        if market_end is None:
            market_end = float(start_ts + 300)

        up_price, down_price = extract_up_down_prices(market)

        state = CycleState(
            slug=slug,
            question=question,
            start_ts=market_start,
            end_ts=market_end,
            up_price=up_price,
            down_price=down_price,
        )
        self.cycles[start_ts] = state
        self._prime_official_ws_slug(state.slug)

        print(
            f"[NOVO CICLO] {state.slug} | start={epoch_to_utc_text(state.start_ts)} | "
            f"end={epoch_to_utc_text(state.end_ts)}"
        )

    def _select_target_sample(self, cycle: CycleState, now_ts: float) -> Tuple[Optional[float], Optional[float], Optional[str]]:
        if self.price_by_ts:
            timestamps = sorted(self.price_by_ts.keys())

            # Preferencia: benchmark em [start, start+12s]
            for ts in timestamps:
                if self.price_src_by_ts.get(ts) != "benchmark":
                    continue
                if cycle.start_ts <= ts <= (cycle.start_ts + 12.0):
                    return ts, self.price_by_ts[ts], self.price_src_by_ts.get(ts)

            # Fallback: qualquer fonte em [start, start+12s]
            for ts in timestamps:
                if cycle.start_ts <= ts <= (cycle.start_ts + 12.0):
                    return ts, self.price_by_ts[ts], self.price_src_by_ts.get(ts)

            # Fallback live depois de 15s de ciclo: pega ponto mais proximo do start em +-30s
            if now_ts >= (cycle.start_ts + 15.0):
                candidate = None
                candidate_dist = None
                candidate_benchmark = None
                candidate_benchmark_dist = None
                for ts in timestamps:
                    if ts < cycle.start_ts or ts > (cycle.start_ts + 30.0):
                        continue
                    dist = abs(ts - cycle.start_ts)
                    if self.price_src_by_ts.get(ts) == "benchmark":
                        if candidate_benchmark is None or dist < candidate_benchmark_dist:
                            candidate_benchmark = ts
                            candidate_benchmark_dist = dist
                    if candidate is None or dist < candidate_dist:
                        candidate = ts
                        candidate_dist = dist
                if candidate_benchmark is not None:
                    return (
                        candidate_benchmark,
                        self.price_by_ts[candidate_benchmark],
                        self.price_src_by_ts.get(candidate_benchmark),
                    )
                if candidate is not None:
                    return candidate, self.price_by_ts[candidate], self.price_src_by_ts.get(candidate)

        # Fallback historico: open do minuto de inicio.
        if now_ts >= (cycle.start_ts + 15.0):
            self.refresh_historical_cache_if_needed()
            h_ts, h_price = self.get_historical_open_at_minute(cycle.start_ts)
            if h_price is not None:
                return h_ts, h_price, "hist.open"

        return None, None, None

    def _select_close_sample(self, cycle: CycleState, now_ts: float) -> Tuple[Optional[float], Optional[float], Optional[str]]:
        decision_deadline = cycle.end_ts + self.decision_delay_seconds

        if self.price_by_ts:
            timestamps = sorted(self.price_by_ts.keys())

            # Preferencia: benchmark em [end, end+delay+0.75s]
            for ts in timestamps:
                if self.price_src_by_ts.get(ts) != "benchmark":
                    continue
                if cycle.end_ts <= ts <= (decision_deadline + 0.75):
                    return ts, self.price_by_ts[ts], self.price_src_by_ts.get(ts)

            # Fallback: qualquer fonte em [end, end+delay+0.75s]
            for ts in timestamps:
                if cycle.end_ts <= ts <= (decision_deadline + 0.75):
                    return ts, self.price_by_ts[ts], self.price_src_by_ts.get(ts)

            # Fallback live estrito: ponto mais proximo do end apenas apos fechamento.
            if now_ts >= (decision_deadline + 0.5):
                candidate = None
                candidate_dist = None
                candidate_benchmark = None
                candidate_benchmark_dist = None
                for ts in timestamps:
                    if ts < cycle.end_ts or ts > (cycle.end_ts + 12.0):
                        continue
                    dist = abs(ts - cycle.end_ts)
                    if self.price_src_by_ts.get(ts) == "benchmark":
                        if candidate_benchmark is None or dist < candidate_benchmark_dist:
                            candidate_benchmark = ts
                            candidate_benchmark_dist = dist
                    if candidate is None or dist < candidate_dist:
                        candidate = ts
                        candidate_dist = dist
                if candidate_benchmark is not None:
                    return (
                        candidate_benchmark,
                        self.price_by_ts[candidate_benchmark],
                        self.price_src_by_ts.get(candidate_benchmark),
                    )
                if candidate is not None:
                    return candidate, self.price_by_ts[candidate], self.price_src_by_ts.get(candidate)

        # Fallback historico estrito: open do minuto do fechamento (forca refresh apos deadline).
        if now_ts >= (decision_deadline + 0.5):
            self.refresh_historical_cache_if_needed(force=True)
            h_ts, h_price = self.get_historical_open_at_minute(cycle.end_ts)
            if h_price is not None and h_ts is not None and h_ts >= cycle.end_ts:
                return h_ts, h_price, "hist.open"

        return None, None, None

    def process_cycle(self, cycle: CycleState, now_ts: float) -> None:
        if cycle.target_price is None:
            t_ts, t_price, t_src = self._select_target_sample(cycle, now_ts)
            if t_price is not None:
                cycle.target_price = t_price
                cycle.target_ts = t_ts
                cycle.target_src = t_src or "unknown"
                print(
                    f"[TARGET] {cycle.slug} | target={cycle.target_price:.8f} | "
                    f"target_ts={epoch_to_utc_text(cycle.target_ts)} | src={cycle.target_src}"
                )
            elif now_ts >= (cycle.start_ts + 15.0) and not cycle.warned_target_delay:
                cycle.warned_target_delay = True
                print(
                    f"[ALERTA] {cycle.slug} sem target capturado ainda (>{15}s apos abertura)."
                )

        if cycle.inferred_result is not None:
            return

        decision_deadline = cycle.end_ts + self.decision_delay_seconds
        if now_ts < decision_deadline:
            return

        official_result = self._maybe_refresh_official_result(cycle, now_ts)
        if official_result in ("UP", "DOWN"):
            cycle.inferred_result = official_result
            if cycle.chainlink_result and cycle.chainlink_result != official_result:
                print(
                    f"[CHECK RESOLUCAO] {cycle.slug} | chainlink={cycle.chainlink_result} | "
                    f"oficial={official_result} | acao=usar_oficial"
                )
        else:
            # Se oficial for obrigatorio sem fallback, aguardamos indefinidamente.
            if self.require_official_resolution and not self.allow_chainlink_fallback:
                if not cycle.warned_resolution_wait:
                    cycle.warned_resolution_wait = True
                    print(
                        f"[AGUARDANDO RESOLUCAO] {cycle.slug} sem resultado oficial ainda; "
                        "modo estrito ativo (sem fallback para Chainlink)."
                    )
                return

            # Se oficial for obrigatorio com fallback, espera a janela extra.
            if self.require_official_resolution and self.allow_chainlink_fallback:
                fallback_deadline = decision_deadline + self.official_resolution_wait_seconds
                if now_ts < fallback_deadline:
                    if not cycle.warned_resolution_wait:
                        cycle.warned_resolution_wait = True
                        print(
                            f"[AGUARDANDO RESOLUCAO] {cycle.slug} sem resultado oficial ainda; "
                            f"aguardando ate {self.official_resolution_wait_seconds:.1f}s de janela extra."
                        )
                    return

            # Caminho por Chainlink (fallback ou modo nao-oficial).
            if cycle.target_price is None:
                return

            if cycle.chainlink_result is None:
                c_ts, c_price, c_src = self._select_close_sample(cycle, now_ts)
                if c_price is None:
                    if not cycle.warned_close_wait:
                        cycle.warned_close_wait = True
                        print(
                            f"[AGUARDANDO] {cycle.slug} aguardando preco final para decisao "
                            f"({self.decision_delay_seconds:.1f}s apos fechamento)."
                        )
                    return

                cycle.close_price = c_price
                cycle.close_ts = c_ts
                cycle.close_src = c_src or "unknown"
                cycle.chainlink_result = "UP" if cycle.close_price >= cycle.target_price else "DOWN"

                chainlink_delta = cycle.close_price - cycle.target_price
                print(
                    f"[INFERENCIA] {cycle.slug} | chainlink={cycle.chainlink_result} | "
                    f"target={cycle.target_price:.8f} ({epoch_to_utc_text(cycle.target_ts)}) | "
                    f"close={cycle.close_price:.8f} ({epoch_to_utc_text(cycle.close_ts)}) | close_src={cycle.close_src} | "
                    f"delta={chainlink_delta:+.8f} | decision_delay={self.decision_delay_seconds:.1f}s"
                )

            if cycle.chainlink_result is None:
                return

            cycle.inferred_result = cycle.chainlink_result
            if self.require_official_resolution and self.allow_chainlink_fallback and not cycle.warned_resolution_fallback:
                cycle.warned_resolution_fallback = True
                print(
                    f"[FALLBACK] {cycle.slug} sem resultado oficial no prazo; "
                    f"usando chainlink={cycle.chainlink_result}."
                )

        result_source = "official" if cycle.official_result else "chainlink"
        delta: Optional[float] = None
        if cycle.close_price is not None and cycle.target_price is not None:
            delta = cycle.close_price - cycle.target_price

        target_text = f"{cycle.target_price:.8f}" if cycle.target_price is not None else "N/A"
        close_text = f"{cycle.close_price:.8f}" if cycle.close_price is not None else "N/A"
        delta_text = f"{delta:+.8f}" if delta is not None else "N/A"
        print(
            f"[DECISAO] {cycle.slug} | resultado={cycle.inferred_result} | fonte={result_source} | "
            f"target={target_text} | close={close_text} | delta={delta_text}"
        )
        try:
            self.result_store.save_inference(
                cycle=cycle,
                delta=delta,
                decision_delay_seconds=self.decision_delay_seconds,
            )
        except Exception as e:
            print(f"[ERRO] Falha ao persistir inferencia: {e}")
        self._settle_paper_trade_if_needed(cycle)

    def prune_old_cycles(self, now_ts: float) -> None:
        # Remove ciclos muito antigos (mais de 30 minutos) para evitar crescer em memoria.
        to_remove = []
        for start_ts, cycle in self.cycles.items():
            if now_ts > (cycle.end_ts + 1800.0):
                to_remove.append(start_ts)
        for key in to_remove:
            self.cycles.pop(key, None)

    def run(self) -> None:
        print(
            f"[INIT] asset={self.asset_name} | stream={self.stream_slug} | feed_id={self.feed_id} | schema={self.schema} | "
            f"abi_index={self.abi_index}"
        )
        print(
            f"[INIT] market_prefix={self.market_prefix} | poll={self.interval_seconds:.2f}s | "
            f"decision_delay={self.decision_delay_seconds:.2f}s"
        )
        print(
            f"[INIT] official_resolution_check=ON | require_official={'ON' if self.require_official_resolution else 'OFF'} | "
            f"official_ws={'ON' if self._ws_enabled else 'OFF'} | "
            f"fallback_chainlink={'ON' if self.allow_chainlink_fallback else 'OFF'} | "
            f"official_wait={self.official_resolution_wait_seconds:.1f}s | poll={self.official_resolution_poll_seconds:.1f}s | "
            f"confirm_polls={self.official_confirmation_polls}"
        )
        print(f"[INIT] csv_output={self.result_store.csv_path}")
        if self.result_store.enable_xlsx and self.result_store.xlsx_path is not None:
            print(f"[INIT] xlsx_output={self.result_store.xlsx_path}")
        print(
            f"[INIT] paper_trade=ON | start_cash={self.paper_starting_cash:.2f} | "
            f"mode={self.paper_sizing_mode} | "
            f"base_shares={self.paper_shares:.2f} | base_stake_usd={self.paper_stake_usd:.4f} | "
            f"sequence={','.join(self.paper_sequence)} | entry_cutoff={self.paper_entry_cutoff_seconds:.1f}s | "
            f"entry_price_range={self.paper_entry_min_price:.2f}-{self.paper_entry_max_price:.2f} | "
            f"gale_multiplier={self.paper_gale_multiplier:.2f} | "
            f"gale_max_steps={'ilimitado' if self.paper_gale_max_steps <= 0 else self.paper_gale_max_steps}"
        )
        print(f"[INIT] paper_csv_output={self.paper_store.csv_path}")
        if self.paper_store.enable_xlsx and self.paper_store.xlsx_path is not None:
            print(f"[INIT] paper_xlsx_output={self.paper_store.xlsx_path}")

        while True:
            loop_start = time.time()
            try:
                now_ts = time.time()
                bucket_start = int(now_ts) - (int(now_ts) % 300)

                # Monitoramos o ciclo atual e o imediatamente anterior.
                for cycle_start in (bucket_start - 300, bucket_start):
                    self.ensure_cycle(cycle_start)

                self.update_live_price_cache()

                for cycle in sorted(self.cycles.values(), key=lambda c: c.start_ts):
                    self.process_cycle(cycle, now_ts)

                self._maybe_open_paper_trade(now_ts)
                self._advance_sequence_for_missed_entries(now_ts)
                self.prune_old_cycles(now_ts)
            except Exception as e:
                print(f"[ERRO] Falha no loop principal: {e}")

            elapsed = time.time() - loop_start
            sleep_for = max(0.05, self.interval_seconds - elapsed)
            time.sleep(sleep_for)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Inferencia rapida de UP/DOWN para mercados 5m (BTC/ETH/SOL) da Polymarket."
    )
    parser.add_argument(
        "--asset",
        default="btc",
        help="Ativo base (btc, eth, sol). Define defaults de stream, prefixo e pasta de relatorios.",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Pasta base de saida dos relatorios (default: Target/<asset> 5min).",
    )
    parser.add_argument(
        "--stream-slug",
        default=None,
        help="Slug do stream da Chainlink (default depende de --asset).",
    )
    parser.add_argument(
        "--market-prefix",
        default=None,
        help="Prefixo do slug do mercado Polymarket (default depende de --asset).",
    )
    parser.add_argument(
        "--poll-interval",
        type=float,
        default=0.60,
        help="Intervalo de polling em segundos (padrao: 0.60).",
    )
    parser.add_argument(
        "--decision-delay",
        type=float,
        default=3.0,
        help="Tempo apos endDate para inferir resultado (padrao: 3.0 segundos).",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=8.0,
        help="Timeout HTTP por request (padrao: 8s).",
    )
    parser.add_argument(
        "--require-official-resolution",
        dest="require_official_resolution",
        action="store_true",
        default=True,
        help="Exige resultado oficial da Polymarket para fechar o ciclo (padrao: ON).",
    )
    parser.add_argument(
        "--no-require-official-resolution",
        dest="require_official_resolution",
        action="store_false",
        help="Permite fechar ciclo sem aguardar resultado oficial (nao recomendado).",
    )
    parser.add_argument(
        "--allow-chainlink-fallback",
        action="store_true",
        help="Permite fallback para Chainlink se o resultado oficial nao chegar no prazo.",
    )
    parser.add_argument(
        "--official-ws",
        dest="official_ws",
        action="store_true",
        default=True,
        help="Usa WebSocket da Polymarket para detectar resultado oficial (padrao: ON).",
    )
    parser.add_argument(
        "--no-official-ws",
        dest="official_ws",
        action="store_false",
        help="Desativa WebSocket oficial e usa apenas polling da Gamma para resultado oficial.",
    )
    parser.add_argument(
        "--official-resolution-wait",
        type=float,
        default=30.0,
        help="Janela extra (s) para aguardar resultado oficial antes de fallback para Chainlink.",
    )
    parser.add_argument(
        "--official-resolution-poll",
        type=float,
        default=2.0,
        help="Intervalo minimo (s) entre tentativas de check do resultado oficial.",
    )
    parser.add_argument(
        "--official-confirmation-polls",
        type=int,
        default=2,
        help="Quantidade de polls consecutivos com mesmo resultado oficial para confirmar fechamento.",
    )
    parser.add_argument(
        "--output-csv",
        default=None,
        help="Arquivo CSV para salvar inferencias em tempo real (default depende de --asset e --output-dir).",
    )
    parser.add_argument(
        "--output-xlsx",
        default=None,
        help="Arquivo Excel (.xlsx) para salvar inferencias em tempo real (default depende de --asset e --output-dir).",
    )
    parser.add_argument(
        "--disable-xlsx",
        action="store_true",
        help="Desativa exportacao XLSX (mantem somente CSV).",
    )
    parser.add_argument(
        "--paper-starting-cash",
        type=float,
        default=100.0,
        help="Banca inicial do paper trade (padrao: 100).",
    )
    parser.add_argument(
        "--paper-shares",
        type=float,
        default=5.0,
        help="Quantidade base de shares por entrada no paper trade (padrao: 5). Ignorado se --paper-stake-usd > 0.",
    )
    parser.add_argument(
        "--paper-stake-usd",
        type=float,
        default=0.0,
        help="Valor base em USD por entrada no paper trade. Se > 0, usa sizing por valor (ex.: 1.0).",
    )
    parser.add_argument(
        "--paper-sequence",
        default="UP,UP,DOWN,DOWN",
        help="Sequencia de lados do paper trade em modo rotativo continuo.",
    )
    parser.add_argument(
        "--paper-entry-cutoff",
        type=float,
        default=20.0,
        help="Janela maxima (segundos apos start do ciclo) para considerar entrada no mercado.",
    )
    parser.add_argument(
        "--paper-entry-min-price",
        type=float,
        default=0.35,
        help="Preco minimo permitido para entrada da posicao no paper trade (padrao: 0.35).",
    )
    parser.add_argument(
        "--paper-entry-max-price",
        type=float,
        default=0.52,
        help="Preco maximo permitido para entrada da posicao no paper trade (padrao: 0.52).",
    )
    parser.add_argument(
        "--paper-gale-multiplier",
        type=float,
        default=2.0,
        help="Multiplicador do gale apos perda (padrao: 2.0).",
    )
    parser.add_argument(
        "--paper-gale-max-steps",
        type=int,
        default=0,
        help="Numero maximo de passos de gale consecutivos (0 = sem limite).",
    )
    parser.add_argument(
        "--paper-output-csv",
        default=None,
        help="Arquivo CSV de saida do paper trade (default depende de --asset e --output-dir).",
    )
    parser.add_argument(
        "--paper-output-xlsx",
        default=None,
        help="Arquivo XLSX de saida do paper trade (default depende de --asset e --output-dir).",
    )
    parser.add_argument(
        "--paper-disable-xlsx",
        action="store_true",
        help="Desativa XLSX do paper trade (mantem somente CSV).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    asset_cfg = resolve_asset_config(args.asset)
    stream_slug = str(args.stream_slug or asset_cfg["stream_slug"])
    market_prefix = str(args.market_prefix or asset_cfg["market_prefix"])
    output_csv, output_xlsx, paper_output_csv, paper_output_xlsx, output_dir = resolve_output_paths(
        asset_cfg=asset_cfg,
        output_dir=args.output_dir,
        output_csv=args.output_csv,
        output_xlsx=args.output_xlsx,
        paper_output_csv=args.paper_output_csv,
        paper_output_xlsx=args.paper_output_xlsx,
    )

    print(
        f"[CONFIG] asset={asset_cfg['asset']} | stream={stream_slug} | "
        f"market_prefix={market_prefix} | output_dir={output_dir}"
    )

    resolver = FastResolver(
        asset_name=asset_cfg["asset"],
        stream_slug=stream_slug,
        market_prefix=market_prefix,
        interval_seconds=max(0.20, args.poll_interval),
        decision_delay_seconds=max(1.0, args.decision_delay),
        timeout_seconds=max(2.0, args.timeout),
        require_official_resolution=bool(args.require_official_resolution),
        allow_chainlink_fallback=args.allow_chainlink_fallback,
        use_official_ws_resolution=bool(args.official_ws),
        official_resolution_wait_seconds=max(0.0, args.official_resolution_wait),
        official_resolution_poll_seconds=max(0.2, args.official_resolution_poll),
        official_confirmation_polls=max(1, args.official_confirmation_polls),
        output_csv=output_csv,
        output_xlsx=output_xlsx,
        enable_xlsx=not args.disable_xlsx,
        paper_starting_cash=args.paper_starting_cash,
        paper_shares=args.paper_shares,
        paper_stake_usd=max(0.0, args.paper_stake_usd),
        paper_sequence=args.paper_sequence,
        paper_entry_cutoff_seconds=max(1.0, args.paper_entry_cutoff),
        paper_entry_min_price=max(0.0, args.paper_entry_min_price),
        paper_entry_max_price=max(0.0, args.paper_entry_max_price),
        paper_gale_multiplier=max(1.0, args.paper_gale_multiplier),
        paper_gale_max_steps=max(0, args.paper_gale_max_steps),
        paper_output_csv=paper_output_csv,
        paper_output_xlsx=paper_output_xlsx,
        enable_paper_xlsx=not args.paper_disable_xlsx,
    )
    resolver.run()


if __name__ == "__main__":
    main()
